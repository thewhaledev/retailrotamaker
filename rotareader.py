from openpyxl import load_workbook
import random
import shutil

workbook = load_workbook(filename="rota.xlsx")

sheet = workbook.active

"""
!Important! This list show the full range of cells in column A, from the
first full-timer to the final retail employee. If more staff are added,
or if the range changes, these cells will also need to be added to the list.
"""

rota_cells = ["A22", "A23", "A24", "A25", "A26", "A27", "A28", "A29", "A30",
"A31", "A32", "A33", "A34", "A35", "A36", "A37", "A38", "A39", "A40", "A41",
"A42", "A43", "A44", "A45", "A46", "A47", "A48", "A49", "A50", "A51", "A52", "A53", "A54",
"A55", "A56", "A57", "A58", "A59", "A60", "A61"]

ignore = ['VACANCY', 'Full-time Staff', 'Part-time Staff', 'Weekend Staff', 'Fixed Term Staff', 'Casual Staff']

cell_values = []
name_cells = []

for item in rota_cells:
    value = sheet[item].value.rstrip()
    cell_values.append(value)
    if value not in ignore:
        name_cells.append(item)

def day_cell_maker(column):
    name_cells = []
    for i in range(len(rota_cells)):
        number = rota_cells[i].strip("A")
        name = cell_values[i]
        if name not in ignore:
            name_cells.append(column + number)
    return name_cells

mon_cells = day_cell_maker("B")
tue_cells = day_cell_maker("F")
wed_cells = day_cell_maker("J")
thu_cells = day_cell_maker("N")
fri_cells = day_cell_maker("R")
sat_cells = day_cell_maker("V")
sun_cells = day_cell_maker("Z")

def employee_dict_maker(day_list):
    """
    There are two values held in this employee dictionary, which makes a dictionary of employees for each day.
    "name" = the name of the employee as found in the rota spreadsheet.
    "time" = the time they come in, which will be used to seperate those on part time hours.
    """
    employees = {}
    for i in range(len(name_cells)):
        if sheet[day_list[i]].value != '-' and sheet[day_list[i]].value != None and sheet[day_list[i]].value != 'AL':
            employees[sheet[name_cells[i]].value] = sheet[day_list[i]].value
    return employees

def all_employees(dict):
    #This function makes a seperate list of the names of all employees
    all_employees = []
    for key in dict.keys():
            all_employees.append(key)
    return all_employees

def part_time_list(dict):
    #This function makes a seperate list of the names of those on part time hours.
    part_timers = []
    for key in dict.keys():
        if dict[key] == 1130 or dict[key] == 1140:
            part_timers.append(key)
    return part_timers

#Now all the functions are called for all the days.

mon_employees = employee_dict_maker(mon_cells)
tue_employees = employee_dict_maker(tue_cells)
wed_employees = employee_dict_maker(wed_cells)
thu_employees = employee_dict_maker(thu_cells)
fri_employees = employee_dict_maker(fri_cells)
sat_employees = employee_dict_maker(sat_cells)
sun_employees = employee_dict_maker(sun_cells)

mon_full_time = all_employees(mon_employees)
tue_full_time = all_employees(tue_employees)
wed_full_time = all_employees(wed_employees)
thu_full_time = all_employees(thu_employees)
fri_full_time = all_employees(fri_employees)
sat_full_time = all_employees(sat_employees)
sun_full_time = all_employees(sun_employees)

mon_part_time  = part_time_list(mon_employees)
tue_part_time = part_time_list(tue_employees)
wed_part_time = part_time_list(wed_employees)
thu_part_time = part_time_list(thu_employees)
fri_part_time = part_time_list(fri_employees)
sat_part_time = part_time_list(sat_employees)
sun_part_time = part_time_list(sun_employees)

"""
!Important! This list holds each shop floor position in rank of their
importance. If needs change, or if shops open and close, this list will
need to be edited.

This is not perfect in its current form, and will require you to change
the rota when lunch covers are on the shop floor, and when full-time
staff are doing lunch covers.
"""

positions = ["M/S, 12PM", "M/S, 1PM", "M/S 2PM", "WL-Ww", "MEZ", "COVER WL-MEZ",
 "WW-WL", "TS", "COVER WW-TS", "M/S, 12PM", "M/S, 1PM", "M/S 2PM", "M/S Extra",
  "M/S Extra", "M/S Extra", "M/S Extra",]

"""
!Important! This list holds the employees who have till "privaleges". This
will have to be changed when such employees change.
"""

important_employees = ["Sylwia Wroblewska", "Gina Lovatt", "Michael Heath",
"Michael Pearson", "Lincoln Keyi", "Richard Grange", "Tinashe Zvobgo",
"Sinuo Guo",]

def position_filler(employee_list, part_time_list):
    """
    This function makes a tuple holding each position and the name of the employee who will fill that position.
    It checks agaisnt three conditions, presented in this order:
    1) That an employee who cannot be on certain positions is not placed in those positions.
    2) That those who are in the part time list are not placed in positions that only full time employees can be.
    3) That there are also employees with "till privaleges" in the main location, where such privalges (refunds etc...) are needed.
    """
    while True:
        random.shuffle(employee_list) # The full time list is shuffled continously until all the proceeding conditions are met.
        if employee_list[3] == "Micheal Heath" or employee_list[5] == "Micheal Heath" or employee_list[6] == "Micheal Heath" or employee_list[8] == "Micheal Heath":
            continue
        if employee_list[3] in part_time_list or employee_list[4] in part_time_list or employee_list[5] in part_time_list or employee_list[6] in part_time_list:
            continue
        if employee_list[0] in important_employees or employee_list[1] in important_employees or employee_list[2] in important_employees:
            break
    day_merge = tuple(zip(positions, employee_list))
    return day_merge

#The function is called here for each day.

monday = position_filler(mon_full_time, mon_part_time)
tuesday = position_filler(tue_full_time, tue_part_time)
wednesday = position_filler(wed_full_time, wed_part_time)
thursday = position_filler(thu_full_time, thu_part_time)
friday = position_filler(fri_full_time, fri_part_time)
saturday = position_filler(sat_full_time, sat_part_time)
sunday = position_filler(sun_full_time, sun_part_time)

#Prints the tuples for the purpose of visualisation and testing.

print(monday)
print(tuesday)
print(wednesday)
print(thursday)
print(friday)
print(saturday)
print(sunday)

#Creates a copy of the rota template, so that this does not need to be done manually.
shutil.copy("daily_rota_template.xlsx", "daily_rotas.xlsx")

#Loads the template as a new workbook.
new_workbook = load_workbook(filename="daily_rotas.xlsx")

#These cells dictate where the names should be written on the template.
position_cells = ["B8", "B10", "B12", "B17", "B18", "B19", "B21",
 "B22", "B23", "B9", "B11", "B13", "B14", "B15", "B16", "B17",]

def daily_rota_maker(sheetname, position_tuple):
    """
    This function places each person from the tuple in their respective cell on the rota sheet.
    The sheetname "if" statements fill in the day of the week on each rota.
    """
    sheet = new_workbook[sheetname]
    for i in range(len(position_tuple)):
        sheet[position_cells[i]] = position_tuple[i][1]
    if sheetname == "Mon":
        sheet["G2"] = "Monday"
    if sheetname == "Tue":
        sheet["G2"] = "Tuesday"
    if sheetname == "Wed":
        sheet["G2"] = "Wednesday"
    if sheetname == "Thu":
        sheet["G2"] = "Thursday"
    if sheetname == "Fri":
        sheet["G2"] = "Friday"
    if sheetname == "Sat":
        sheet["G2"] = "Saturday"
    if sheetname == "Sun":
        sheet["G2"] = "Sunday"
    new_workbook.save("daily_rotas.xlsx")

#The function is called again for each day of the week.

daily_rota_maker("Mon", monday)
daily_rota_maker("Tue", tuesday)
daily_rota_maker("Wed", wednesday)
daily_rota_maker("Thu", thursday)
daily_rota_maker("Fri", friday)
daily_rota_maker("Sat", saturday)
daily_rota_maker("Sun", sunday)

print(mon_full_time)
print(mon_part_time)

input("Press ENTER to exit")
